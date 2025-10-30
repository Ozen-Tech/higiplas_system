from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, and_
from datetime import datetime, timedelta
from typing import Optional, Literal
import io
import json

from ..db import models
from ..core.logger import stock_operations_logger

ExportFormat = Literal['json', 'pdf', 'xlsx']

class StockReportService:
    
    @staticmethod
    def generate_weekly_stock_report(
        db: Session,
        empresa_id: int,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        format: ExportFormat = "json"
    ) -> dict:
        if not end_date:
            end_date = datetime.now()
        if not start_date:
            start_date = end_date - timedelta(days=7)
        
        stock_operations_logger.info(
            f"[generateWeeklyReport] START - empresa_id={empresa_id}, "
            f"start_date={start_date}, end_date={end_date}, format={format}"
        )
        
        movimentacoes = db.query(models.MovimentacaoEstoque).join(
            models.Produto
        ).filter(
            and_(
                models.Produto.empresa_id == empresa_id,
                models.MovimentacaoEstoque.data_movimentacao >= start_date,
                models.MovimentacaoEstoque.data_movimentacao <= end_date
            )
        ).options(
            joinedload(models.MovimentacaoEstoque.produto),
            joinedload(models.MovimentacaoEstoque.usuario)
        ).order_by(
            models.MovimentacaoEstoque.data_movimentacao.desc()
        ).all()
        
        total_movimentacoes = len(movimentacoes)
        entradas = sum(1 for m in movimentacoes if m.tipo_movimentacao == 'ENTRADA')
        saidas = sum(1 for m in movimentacoes if m.tipo_movimentacao == 'SAIDA')
        
        detalhes = []
        for mov in movimentacoes:
            detalhes.append({
                "id": mov.id,
                "data": mov.data_movimentacao.isoformat() if mov.data_movimentacao else None,
                "produto": mov.produto.nome if mov.produto else "Produto não encontrado",
                "codigo_produto": mov.produto.codigo if mov.produto else None,
                "tipo": mov.tipo_movimentacao,
                "quantidade": mov.quantidade,
                "quantidade_antes": mov.quantidade_antes,
                "quantidade_depois": mov.quantidade_depois,
                "usuario": mov.usuario.nome if mov.usuario else "Usuário não encontrado",
                "origem": mov.origem,
                "observacao": mov.observacao
            })
        
        report_data = {
            "periodo": f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            "total_movimentacoes": total_movimentacoes,
            "resumo": {
                "entradas": entradas,
                "saidas": saidas
            },
            "detalhes": detalhes
        }
        
        stock_operations_logger.info(
            f"[generateWeeklyReport] SUCCESS - empresa_id={empresa_id}, "
            f"total_movimentacoes={total_movimentacoes}, entradas={entradas}, saidas={saidas}"
        )
        
        if format == "json":
            return report_data
        elif format == "pdf":
            return StockReportService._export_to_pdf(report_data)
        elif format == "xlsx":
            return StockReportService._export_to_xlsx(report_data)
        else:
            return report_data
    
    @staticmethod
    def _export_to_pdf(report_data: dict) -> dict:
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import cm
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=1
            )
            
            title = Paragraph(f"Relatório Semanal de Estoque<br/>{report_data['periodo']}", title_style)
            elements.append(title)
            
            resumo_data = [
                ['Total de Movimentações', 'Entradas', 'Saídas'],
                [
                    str(report_data['total_movimentacoes']),
                    str(report_data['resumo']['entradas']),
                    str(report_data['resumo']['saidas'])
                ]
            ]
            
            resumo_table = Table(resumo_data, colWidths=[6*cm, 6*cm, 6*cm])
            resumo_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(resumo_table)
            elements.append(Spacer(1, 20))
            
            detail_title = Paragraph("Detalhes das Movimentações", styles['Heading2'])
            elements.append(detail_title)
            elements.append(Spacer(1, 10))
            
            detail_data = [['Data', 'Produto', 'Tipo', 'Qtd', 'Antes', 'Depois', 'Usuário', 'Origem']]
            
            for item in report_data['detalhes'][:50]:
                data_formatada = datetime.fromisoformat(item['data']).strftime('%d/%m %H:%M') if item['data'] else '-'
                detail_data.append([
                    data_formatada,
                    item['produto'][:20] + '...' if len(item['produto']) > 20 else item['produto'],
                    item['tipo'],
                    str(item['quantidade']),
                    str(item['quantidade_antes']) if item['quantidade_antes'] is not None else '-',
                    str(item['quantidade_depois']) if item['quantidade_depois'] is not None else '-',
                    item['usuario'][:15] + '...' if len(item['usuario']) > 15 else item['usuario'],
                    item['origem'][:10] if item['origem'] else '-'
                ])
            
            detail_table = Table(detail_data, colWidths=[3*cm, 5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 3.5*cm, 2.5*cm])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            elements.append(detail_table)
            
            doc.build(elements)
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            import base64
            pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
            
            return {
                "format": "pdf",
                "filename": f"relatorio_estoque_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                "content": pdf_base64,
                "content_type": "application/pdf"
            }
            
        except ImportError:
            stock_operations_logger.warning("[exportToPDF] reportlab not installed, returning JSON")
            return {"error": "PDF export requires reportlab library", "data": report_data}
        except Exception as e:
            stock_operations_logger.error(f"[exportToPDF] ERROR - {str(e)}")
            return {"error": f"Failed to generate PDF: {str(e)}", "data": report_data}
    
    @staticmethod
    def _export_to_xlsx(report_data: dict) -> dict:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório de Estoque"
            
            ws['A1'] = "RELATÓRIO SEMANAL DE ESTOQUE"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = report_data['periodo']
            ws['A2'].font = Font(size=12)
            
            ws['A4'] = "Total de Movimentações"
            ws['B4'] = report_data['total_movimentacoes']
            ws['A5'] = "Entradas"
            ws['B5'] = report_data['resumo']['entradas']
            ws['A6'] = "Saídas"
            ws['B6'] = report_data['resumo']['saidas']
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            headers = ['Data', 'Produto', 'Código', 'Tipo', 'Quantidade', 'Qtd Antes', 'Qtd Depois', 'Usuário', 'Origem', 'Observação']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=8, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            for row_idx, item in enumerate(report_data['detalhes'], start=9):
                data_formatada = datetime.fromisoformat(item['data']).strftime('%d/%m/%Y %H:%M') if item['data'] else '-'
                ws.cell(row=row_idx, column=1, value=data_formatada)
                ws.cell(row=row_idx, column=2, value=item['produto'])
                ws.cell(row=row_idx, column=3, value=item['codigo_produto'])
                ws.cell(row=row_idx, column=4, value=item['tipo'])
                ws.cell(row=row_idx, column=5, value=item['quantidade'])
                ws.cell(row=row_idx, column=6, value=item['quantidade_antes'])
                ws.cell(row=row_idx, column=7, value=item['quantidade_depois'])
                ws.cell(row=row_idx, column=8, value=item['usuario'])
                ws.cell(row=row_idx, column=9, value=item['origem'])
                ws.cell(row=row_idx, column=10, value=item['observacao'])
            
            for col in range(1, 11):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            buffer = io.BytesIO()
            wb.save(buffer)
            xlsx_bytes = buffer.getvalue()
            buffer.close()
            
            import base64
            xlsx_base64 = base64.b64encode(xlsx_bytes).decode('utf-8')
            
            return {
                "format": "xlsx",
                "filename": f"relatorio_estoque_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "content": xlsx_base64,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
            
        except ImportError:
            stock_operations_logger.warning("[exportToXLSX] openpyxl not installed, returning JSON")
            return {"error": "XLSX export requires openpyxl library", "data": report_data}
        except Exception as e:
            stock_operations_logger.error(f"[exportToXLSX] ERROR - {str(e)}")
            return {"error": f"Failed to generate XLSX: {str(e)}", "data": report_data}
